import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def convert_csv_to_excel(csv_file, excel_file):
    # Baca file CSV
    df = pd.read_csv(csv_file, sep=';')

    
    # Simpan sebagai file Excel
    df.to_excel(excel_file, index=False)

    # Mengatur lebar kolom dan format
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Atur font dan format untuk baris pertama (header)
    header_font = Font(size=16, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4c7e4c", end_color="4c7e4c", fill_type="solid")  # Latar belakang gelap

    # Terapkan font dan latar belakang untuk baris pertama
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')  # Rata kiri untuk header

    # Atur font ukuran 14 untuk seluruh sheet, kecuali baris pertama
    content_font = Font(size=14)

    for row in sheet.iter_rows(min_row=2):  # Mulai dari baris ke-2
        for cell in row:
            cell.font = content_font
            cell.alignment = Alignment(horizontal='left', vertical='center')  # Rata kiri untuk konten

    # Mengatur lebar kolom (10 unit lebih dari panjang nilai)
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 10  # menambahkan 10 unit ekstra
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(excel_file)
    print(f"File Excel berhasil disimpan di: {excel_file}")

if __name__ == "__main__":
    # Lokasi file CSV dari argument pertama
    csv_file = sys.argv[1]
    
    # Lokasi file Excel yang akan dibuat
    excel_file = csv_file.replace(".csv", ".xlsx")
    
    # Konversi CSV ke Excel
    convert_csv_to_excel(csv_file, excel_file)
